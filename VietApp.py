from tkinter import *
from tkinter import ttk, messagebox

from PIL import ImageTk, Image

from google_trans_new import google_translator
from deep_translator import GoogleTranslator
from bing_image_downloader.downloader import download

from bs4 import BeautifulSoup

import requests
import subprocess as sp
import tempfile
import sys, os
import random
import re
from openpyxl import Workbook, load_workbook

translator = google_translator()
global script_dir

#Sets the functionality of running the .exe or the .py
if getattr(sys, 'frozen', False):
    script_dir = os.path.dirname(sys.executable)
elif __file__:
    script_dir = os.path.dirname(os.path.abspath(__file__))

#define search parameters, can change within functions
_search_params = {
    'query_search' : 'default', #Will change
    'limit' : 1,
    'output_dir': 'Images',
    'filter': 'photo',
    'verbose': False
}

#save_word
#Pre: Inputs an english word with a vietnamese word
#Post: append the word to the excel sheet
def save_word(word, viet):
    path = os.path.join(script_dir, 'Review')

    if os.path.exists(path) == False: #Create a new Review Folder
        os.mkdir(path)

    os.chdir(rf"{path}") #IN the review tab

    path = os.path.join(path, 'Review.xlsx')

    if os.path.exists(path) == False:
        workbook = Workbook()
        workbook.save(path)  #Will create a NEW workbook
        sheet = workbook.active
        defaultData = [["Viet Word", "English Word"], [viet, word]]
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        for row in defaultData:
            sheet.append(row)

    else:
        workbook = load_workbook(filename='Review.xlsx')
        wordData = [[viet, word]]
        sheet = workbook.active
        if(check_duplicate(word, sheet) == -1):
            for row in wordData:
                sheet.append(row)

        #else don't append

    workbook.save(path) #Will save the localWorkBook
    os.chdir("..")

#translate_word
#Pre: Input an english word
#Post: Change the global variable viet_word to contain the translated word
def translate_word(word):
    try:
        viet_word = translator.translate(word, lang_tgt='vi', lang_src='en')
    except:
        try:
            viet_word = GoogleTranslator(source="auto", target="vi").translate(text=word)
        except:
            viet_word = CURR_WORD

    return viet_word

#add_viet_word
#Pre: Input an english word
#Post: Output to the random screen the vietnamese text
def add_viet_word(word):
    global viet_label
    viet_frame = ttk.Frame(SCREEN_RAND)
    viet_frame.grid(row=0, column=1)

    global viet_word_to_save
    viet_word_to_save = translate_word(word)

    viet_label.destroy() #destroy previous

    viet_label = ttk.Label(viet_frame, text=viet_word_to_save, font=('Helvetica bold', 20))
    viet_label.grid(row=0,column=0)

#give_up
#Pre: None
#Post: A messagebox on whether or not to give up and save the word
def give_up():
    if messagebox.askyesno(message=f"Do you really want to give up?", title="Don't give up!") == True:
        if messagebox.askquestion(message=f"That's okay! The word is \"{CURR_WORD}\". Do you want to save?", title="'Til Next Time!") == "yes":
            save_word(CURR_WORD, viet_word_to_save)
        get_random_word(WORDS_LIST)
    else:
        messagebox.showwarning(message="You got this!", title="Keep going!")

def get_hint(word):
    #search up the definition and classifer (noun, verb, adjective, etc.)
    word_site = f'https://api.dictionaryapi.dev/api/v2/entries/en/{word}'
    response = requests.get(word_site)
    content = response.json()
    for cont in content:
        for i in range(len(cont['meanings'])):
            dict = cont['meanings'][i] #get the dict
            defineDict = dict["definitions"][0]
            definition = defineDict["definition"]

    messagebox.showinfo(message=f"Definition: {definition}", title="Definition")

#open_menu
#Pre: Input to a ttk screen
#Post: Exit out of the current screen and return to main screen
def open_menu(screen):
    exit_out(screen)

#check_correct
#Pre: The user inputed word and the current word (currW) that is to be guessed
#Post: Returns a success or failure through a messagebox
def check_correct(word, currW):
    #clear entry field
    ENTRY_TEXT.delete(0, END)
    word = word.lower() #lowercase
    word = word.strip() #no whitespace

    if word == currW: #the values are the same
        if messagebox.askquestion(message=f"That is the correct word! Word is \"{CURR_WORD}\". Would you like to save?", title="Correct") == "yes":
            save_word(CURR_WORD, viet_word_to_save)
        get_random_word(WORDS_LIST)
    else:
        messagebox.showwarning(message="That is not the correct word! Try again.", title="Wrong")

#get_random_word
#Pre: A list of words
#Post: Picks a random word and downloads an image, adds the image to the GUI, and add the viet word
def get_random_word(words):
    random_word = words[random.randint(0,2999)]
    global CURR_WORD
    CURR_WORD = '' + random_word #put in global current word a copy

    #insert check_duplicate here

    add_image(random_word) #Add the word into the Images Folder
    get_image(random_word, SCREEN_P)
    add_viet_word(random_word)

#add_image
#Pre: Input an english word
#Post: Downloads an image in the local "Images" Folder
def add_image(word):
    fileName = word
    _search_params['query_search'] = word

    if os.path.exists(os.path.join(script_dir, "Images", fileName)) == True: #If there is already an instance of this word, don't download
         return

    try:
        download(_search_params['query_search'],  _search_params['limit'], _search_params['output_dir'], _search_params['filter'], _search_params['verbose'])

    except:
        print("Error DOWNLOADING!")

#get_image
#Pre: Inputs an english word, and the screen it wants to output to
#Post: Retrieves the specific image within the "Images" Folder to output
def get_image(word, screen):
    # script_dir = os.path.dirname(os.path.abspath(__file__)) #<-- absolute dir the script is in
    path = os.path.join(script_dir, "Images", word) #Get the first result

    if(os.path.exists(path) == False): #Adds image if it doesn't already exist for Review Tab
        add_image(word)

    #definitely a temp fix, optimize later on
    for root_, dirs_, files in os.walk(path):
        for file in files:
            path = os.path.join(path, file)

    image_ = Image.open(path)

    n_image = image_.resize((320, 320))

    photo = ImageTk.PhotoImage(n_image)
    random_image = ttk.Label(screen, image=photo)
    random_image.image = photo #contain another reference
    random_image.grid(row=0, column=0, sticky="nwes")


#initialize_list
#Pre: None
#Post: Given a specific word site, web scrapes it for the 3000 words, returns a list of words
def initalize_list():

    word_site = "https://www.ef.edu/english-resources/english-vocabulary/top-3000-words/"

    response = requests.get(word_site)
    soup = BeautifulSoup(response.content, "html.parser") #main overhead
    results = soup.find(id="main-content")
    container = results.find("div", class_="field-item even")

    global WORDS_LIST

    WORDS_LIST = container.find("p")
    for a in soup.findAll('br'):
        a.extract()

    WORDS_LIST = list(WORDS_LIST)
    WORDS_LIST = [str(x) for x in WORDS_LIST] #convert to string as list comprehension

#handler
#Pre: event namespace
#Post: Used when pressing enter, check to see if the user's input is correct
def handler(event):
    check_correct(USER_INPUT.get(), CURR_WORD)

#learn_random
#Pre: None
#Post: Outputs the necessary widgets of learning a random word
def learn_random():
    global SCREEN_RAND
    global SCREEN_P
    SCREEN_RAND = Toplevel(mainScreen)
    SCREEN_RAND.geometry("600x600")
    SCREEN_RAND.rowconfigure([0, 1, 2], weight="1", minsize="5")
    SCREEN_RAND.columnconfigure([0, 1, 2], weight="1", minsize="5")
    SCREEN_RAND.resizable(FALSE, FALSE)

    SCREEN_P = ttk.Frame(SCREEN_RAND, relief= RIDGE, borderwidth=5) #main frame for images
    SCREEN_P.grid(row=1, column=1)

    root.state('withdrawn')

    #Exit Button
    exit_frame = ttk.Frame(SCREEN_RAND, relief=RAISED, borderwidth=1)
    exit_frame.grid(row=0, column=0)


    exit_button = ttk.Button(master=exit_frame, text="Exit", command= lambda: open_menu(SCREEN_RAND))
    exit_button.grid(row=0, column=0)

    #info button

    info_frame = ttk.Frame(SCREEN_RAND, relief=RAISED, borderwidth=1)
    info_frame.grid(row=0, column=2)

    info_button = ttk.Button(master=info_frame, text="Info", command= lambda: display_info("random_info"))
    info_button.grid(row=0, column=0)

    #Lower Entries

    screen_lower_buttons = ttk.Labelframe(SCREEN_RAND, text="Entries", relief=RAISED)
    screen_lower_buttons.grid(row=2, column=1)


    hint_button = ttk.Button(master=screen_lower_buttons, text="Hint", command= lambda: get_hint(CURR_WORD))
    hint_button.grid(row=1, column=0, padx=5, pady=5, ipady=5, ipadx=5)

    giveup_button = ttk.Button(master=screen_lower_buttons, text="Give up", command= lambda: give_up())
    giveup_button.grid(row=1, column=1, pady=5, ipady=5, ipadx=5)

    global USER_INPUT
    global ENTRY_TEXT

    USER_INPUT = StringVar()
    ENTRY_TEXT = ttk.Entry(master=screen_lower_buttons, width=20, textvariable=USER_INPUT)
    ENTRY_TEXT.grid(row=0, column=1, pady=10, ipady=5, ipadx=5)


    enter_button = ttk.Button(master=screen_lower_buttons, text="Enter", command= lambda: check_correct(
        USER_INPUT.get(), CURR_WORD))
    enter_button.grid(row=1, column=2,padx=5, pady=5, ipady=5, ipadx=5)

    get_random_word(WORDS_LIST)
    SCREEN_RAND.bind('<Return>', handler)

#handler_enter_word
#Pre: event namespace
#Post: Used for the review_tab when trying to locate a specific word in the Excel
def handler_enter_word(event):
    enter_word(NEW_WORD.get())

#random_review
#Pre: None
#Post: Outputs the necessary widgets of reviewing learned words
def random_review():

    path = os.path.join(script_dir, 'Review')
    os.chdir(rf"{path}") #IN the review tab

    workbook = load_workbook(filename='Review.xlsx')
    sheet = workbook.active

    num_rows = sheet.max_row #exclude the headers
    random_row_num = random.randint(2, num_rows) #inclusive

    right_button = ttk.Button(lower_buttons_frame, text=">>", command=lambda : forward_row(random_row_num+1))
    left_button = ttk.Button(lower_buttons_frame, text="<<", command=lambda  : backward_row(random_row_num-1))

    if(random_row_num is sheet.max_row):
        #set the button to disable
        right_button = ttk.Button(lower_buttons_frame, text=">>", state=DISABLED)
    if(random_row_num == 2):
        #set the button to disable
        left_button = ttk.Button(lower_buttons_frame, text="<<", state=DISABLED)



    right_button.grid(row=1, column=4)
    left_button.grid(row=1, column=0)

    viet_word = sheet.cell(row=random_row_num, column=1).value
    word = sheet.cell(row=random_row_num, column=2).value

    load_notebook(viet_word, word)
    path = os.path.join(path, 'Review.xlsx')
    workbook.save(path) #Will save the localWorkBook
    os.chdir("..")


#load_notebook
#Pre: Vietnamese word, english word
#Post: Modfies the Notebook widget in the Review Tab for the Picture, Word, English, and Viet Tabs
def load_notebook(viet_word, word):

    global define
    global eng_sent

    get_image(word, Picture_Frame) #load image onto the Picture Frame
    word_site = f'https://api.dictionaryapi.dev/api/v2/entries/en/{word}'
    response = requests.get(word_site)
    content = response.json()
    content = content[0]
    for i in range(len(content['meanings'])):
        dict = content['meanings'][i] #get the dict
        for j in range(len(dict["definitions"])): #length of [] of definitions
            defineDict = dict["definitions"][j] #iterate through the array of definitions, outputting a dict
            define = defineDict["definition"]
            if "example" in defineDict:
                try:
                    eng_sent = defineDict["example"]
                except:
                    eng_sent = "There is no example."
                break
            else:
                eng_sent = "There is no example." #Could set to a definition?
        #break out of the very FIRST example, getting the additional definition

    global define_label
    global example_label
    global define_label_v
    global example_label_v
    global eng_define_label
    global viet_define_label
    global eng_sent_label
    global viet_sent_label
    global word_label
    global viet_label


    define_label.destroy()
    example_label.destroy()

    define_label_v.destroy()
    example_label_v.destroy()

    word_label.destroy()
    viet_label.destroy()

    eng_define_label.destroy()
    viet_define_label.destroy()

    eng_sent_label.destroy()
    viet_sent_label.destroy()

    viet_sent = translate_word(eng_sent) #outputs a viet sentence
    viet_define = translate_word(define)

    #Header labels
    define_label = ttk.Label(En_Frame, text="Define: ",font=('Helvetica bold', 15), justify=LEFT)
    example_label = ttk.Label(En_Frame, text="Example: ", font=('Helvetica bold', 15),  justify=LEFT)


    define_label.grid(row=0, column=0, pady=5)
    example_label.grid(row=2, column=0, pady=5, padx=5)


    define_label_v = ttk.Label(Viet_Frame, text="Define: ",font=('Helvetica bold', 15),  justify=LEFT)
    example_label_v = ttk.Label(Viet_Frame, text="Example: ", font=('Helvetica bold', 15), justify=LEFT)

    define_label_v.grid(row=0, column=0, pady=5)
    example_label_v.grid(row=2, column=0, pady=5)

    word_label = ttk.Label(Word_Frame, text=word, font=('Helvetica bold', 15))
    viet_label = ttk.Label(Word_Frame, text=viet_word, font=('Helvectica bold', 15))

    word_label.place(relx=0.5, rely=0.3, anchor=CENTER)
    viet_label.place(relx=0.5, rely=0.7, anchor=CENTER)

    #end of headers

    eng_define_label = ttk.Label(En_Frame, text=define, font=('Helvetica bold', 12),wraplength=190, justify=LEFT)
    viet_define_label = ttk.Label(Viet_Frame, text=viet_define, font=('Helvetica bold', 12), wraplength=190, justify=LEFT)


    eng_define_label.grid(row=1, column=1)
    viet_define_label.grid(row=1, column=1)

    eng_sent_label = ttk.Label(En_Frame, text=eng_sent, font=('Helvetica bold', 12), wraplength=190, justify=LEFT)
    viet_sent_label = ttk.Label(Viet_Frame, text = viet_sent, font=('Helvetica bold', 12), wraplength=190, justify=LEFT)


    eng_sent_label.grid(row=3, column=1)
    viet_sent_label.grid(row=3, column=1)

#forward_row
#Pre: current row number of Excel
#Post: Updates the forward button on the review tab to move to the next entry of Excel
def forward_row(row):
    # script_dir = os.path.dirname(os.path.abspath(__file__)) #absolute directory of the script
    path = os.path.join(script_dir, 'Review')
    os.chdir(rf"{path}") #IN the review tab
    workbook = load_workbook(filename='Review.xlsx')
    sheet = workbook.active


    right_button = ttk.Button(lower_buttons_frame, text=">>", command=lambda : forward_row(row+1))
    left_button = ttk.Button(lower_buttons_frame, text="<<", command=lambda  : backward_row(row-1))

    if(row >= sheet.max_row):
        #set the button to disable
        right_button = ttk.Button(lower_buttons_frame, text=">>", state=DISABLED)

    viet_word = sheet.cell(row=row, column=1).value #get the next word
    word = sheet.cell(row=row, column=2).value #get the english word
    #word will be viet_t

    right_button.grid(row=1, column=4)
    left_button.grid(row=1, column=0)
    path = os.path.join(path, 'Review.xlsx')
    workbook.save(path) #Will save the localWorkBook
    os.chdir("..")

    load_notebook(viet_word, word)

#backword_row
#Pre: current row number of Excel
#Post: Updates the backwards button on the review tab to move to the previous entry of Excel
def backward_row(row):
    # script_dir = os.path.dirname(os.path.abspath(__file__)) #absolute directory of the script
    path = os.path.join(script_dir, 'Review')
    os.chdir(rf"{path}") #IN the review tab
    workbook = load_workbook(filename='Review.xlsx')
    sheet = workbook.active


    right_button = ttk.Button(lower_buttons_frame, text=">>", command=lambda : forward_row(row+1))
    left_button = ttk.Button(lower_buttons_frame, text="<<", command=lambda  : backward_row(row-1))


    if(row == 2):
        #set the button to disable
        left_button = ttk.Button(lower_buttons_frame, text="<<", state=DISABLED)


    viet_word = sheet.cell(row=row, column=1).value
    word = sheet.cell(row=row, column=2).value

    right_button.grid(row=1, column=4)
    left_button.grid(row=1, column=0)
    path = os.path.join(path, 'Review.xlsx')
    workbook.save(path) #Will save the localWorkBook
    os.chdir("..")

    load_notebook(viet_word, word)

#check_dupliacate
#Pre: A english word, the excel workbook
#Post: Returns the row index of the word. If not found, return a -1 (indicates a new word)
def check_duplicate(word, sheet):
    row=2
    found = False
    for value in sheet.iter_rows(min_row=2, values_only=True):

        if(value[1] == word):
            found = True
            break

        row += 1


    if(found):
        return row
    else:
        return -1

#enter_word
#Pre: An english word
#Post: Appends the Excel Sheet with the corresponding english and vietnamese word
def enter_word(word):
    #append to the worksheet what word you want, and pass it through load_notebook
    # script_dir = os.path.dirname(os.path.abspath(__file__)) #absolute directory of the script
    path = os.path.join(script_dir, 'Review')

    NEW_INPUT.delete(0, END)

    #lower_case for word
    #try and except here for irregular word

    os.chdir(rf"{path}") #IN the review tab
    workbook = load_workbook(filename="Review.xlsx")
    sheet = workbook.active

    #a function to not add the word to the excel if it already exists
    #No duplicates, add the word
    viet_word = translate_word(word)

    row_num = check_duplicate(word, sheet)
    #completely new word (-1 is new word)
    if(row_num == -1):
        new_data = [[viet_word, word]]
        for row in new_data:
            sheet.append(row) #Put new word into Review.xlsx

        #update the buttons
        right_button = ttk.Button(lower_buttons_frame, text=">>", state=DISABLED) #Move to the end of the list
        left_button = ttk.Button(lower_buttons_frame, text="<<", command=lambda  : backward_row(sheet.max_row-1))
        right_button.grid(row=1, column=4)
        left_button.grid(row=1, column=0)

    else:

        right_button = ttk.Button(lower_buttons_frame, text=">>", command= lambda :forward_row(row_num+1))
        left_button = ttk.Button(lower_buttons_frame, text="<<", command= lambda  : backward_row(row_num-1))

        if(row_num is sheet.max_row):
            #set the button to disable
            right_button = ttk.Button(lower_buttons_frame, text=">>", state=DISABLED)

        if(row_num == 2):
            #set the button to disable
            left_button = ttk.Button(lower_buttons_frame, text="<<", state=DISABLED)

        right_button.grid(row=1, column=4)
        left_button.grid(row=1, column=0)

    path = os.path.join(path, "Review.xlsx")
    workbook.save(path) #Will save the localWorkBook
    os.chdir("..")

    load_notebook(viet_word, word)  #load the notebook after chdir


#open_excel
#Pre: None
#Post: Opens a notepad .txt file that contains the Excel Sheet as a temporary file
def open_excel():
    path = os.path.join(script_dir, 'Review')
    os.chdir(rf"{path}")

    workbook = load_workbook(filename='Review.xlsx')
    sheet = workbook.active

    os.chdir("..")

    with  tempfile.NamedTemporaryFile(mode="w+t", encoding="utf-8", newline='\n', delete=False) as f:
        for row in sheet.iter_rows(min_row=1, values_only=True):
            x = 25 - len(row[0])
            f.write(f"{row[0]}{' ' * x}{row[1]} \n")

        f.flush()

        programName = 'notepad.exe'
        fileName = f.name
        process = sp.Popen([programName, fileName])

        f.close()

#translate_individual
#Pre: None
#Post: At the current position of the Review Tab, translate the individual
#words of the english examples and definitions word by word. This reinforces
#vocabularly at an individual level

#Within the english sentence, for however many words there are, set them equal to a viet word
    #e.g. "There is no example"
    # "There" = Kia
    # "is" = La
    # 'no' = khong
    # 'example' = vi du
def translate_individual():
    #from the english definition and example
    #define and eng_sent
    define_list = re.sub(pattern=r'[^\w\s]', repl="", string=define)
    eng_list = re.sub(pattern=r'[^\w\s]', repl="", string=eng_sent)

    define_list = define_list.split()
    eng_list = eng_list.split()

    with tempfile.NamedTemporaryFile(mode="w+t", encoding="utf-8", newline='\n', delete=False) as f:
        f.write("Definition Individual Translation\n\n")
        for word in define_list:
            viet_translated = translate_word(word)
            x = 20 - len(word)
            f.write(f"{word}{' ' * x}{viet_translated}\n")

        f.write("\n\n")
        f.write("Example Individual Translation\n\n")

        for word in eng_list:
            viet_translated = translate_word(word)
            x = 20 - len(word)
            f.write(f"{word}{' ' * x}{viet_translated}\n")

        f.flush()

        programName = 'notepad.exe'
        fileName = f.name
        process = sp.Popen([programName, fileName])

        f.close()

#review_words
#Pre: None
#Post: Outputs the necessary widgets to create the Reviews windows
def review_words():
    global REVIEW_SCREEN
    REVIEW_SCREEN = Toplevel(mainScreen)
    REVIEW_SCREEN.geometry("600x600")
    REVIEW_SCREEN.rowconfigure([1, 2, 3], weight=1, minsize=5)
    REVIEW_SCREEN.columnconfigure([1, 2, 3], weight=1, minsize=5)
    REVIEW_SCREEN.resizable(FALSE, FALSE)
    root.state('withdrawn')

    REVIEW_N = ttk.Frame(REVIEW_SCREEN)
    REVIEW_N.grid(row=1, column=1)

    review_notebook = ttk.Notebook(REVIEW_N, width=320)
    review_notebook.grid(row=0, column=0, padx=25)

    global Picture_Frame
    global Word_Frame
    global En_Frame
    global Viet_Frame

    Picture_Frame = ttk.Frame(review_notebook)   # first page, which would get widgets gridded into it
    Word_Frame = ttk.Frame(review_notebook)
    En_Frame = ttk.Frame(review_notebook)   # second page
    Viet_Frame = ttk.Frame(review_notebook) # Third page


    review_notebook.add(Picture_Frame, text='Picture')
    review_notebook.add(Word_Frame, text='Word')
    review_notebook.add(En_Frame, text='English')
    review_notebook.add(Viet_Frame, text="Viet")


    #Exit Button
    exit_frame = ttk.Frame(REVIEW_SCREEN, relief=RAISED, borderwidth=1)
    exit_frame.grid(row=0, column=0, padx=20, pady=20)

    exit_button = ttk.Button(master=exit_frame, text="Exit", command= lambda: open_menu(REVIEW_SCREEN))
    exit_button.grid(row=0, column=0)

    #info button

    info_frame = ttk.Frame(REVIEW_SCREEN, relief=RAISED, borderwidth=1)
    info_frame.grid(row=0, column=2)

    info_button = ttk.Button(master=info_frame, text="Info", command= lambda: display_info("review_info"))
    info_button.grid(row=0, column=0)

    #random button
    random_frame = ttk.Frame(REVIEW_SCREEN)
    random_frame.grid(row=0, column=1)

    random_button = ttk.Button(random_frame, text="Random Word", command= lambda: random_review())
    random_button.grid(row=0,column=0)


    #Lower Entries: Buttons <<, >>, entry for input a word, a random button to take from excel
    #Could do additional buttons to go to the beginning and go to the end!

    #create workbook if it hasn't already been created by Random_words
    # script_dir = os.path.dirname(os.path.abspath(__file__)) #absolute directory of the script
    path = os.path.join(script_dir, 'Review')

    if(os.path.exists(path)== False):
        os.mkdir(path) #Create a Review Tab

    os.chdir(rf"{path}") #IN the review tab
    path = os.path.join(path, 'Review.xlsx')


    if os.path.exists(path) == False:
        workbook = Workbook()
        workbook.save(path)  #Will create a NEW workbook
        sheet = workbook.active
        defaultData = [["Viet Word", "English Word"], ["xin chào", "hello"]]
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        for row in defaultData:
            sheet.append(row)
        workbook.save(path)


    #load the first entry
    workbook = load_workbook(filename="Review.xlsx")
    sheet = workbook.active
    max_row = sheet.max_row
    viet_word = sheet.cell(row=2, column=1).value
    word = sheet.cell(row=2, column=2).value
    os.chdir("..")

    load_notebook(viet_word, word)

    global lower_buttons_frame
    lower_buttons_frame = ttk.Labelframe(REVIEW_SCREEN, text="Entries")
    lower_buttons_frame.grid(row=2, column=0, rowspan=3, columnspan=5)

    input_dir = ttk.Label(lower_buttons_frame,  text="Enter a word to create a new entry OR go to that word.")
    input_dir.grid(row=0, column=2, pady=5)

    global right_button
    global left_button


    global NEW_WORD
    global NEW_INPUT

    NEW_WORD = StringVar()
    NEW_INPUT = ttk.Entry(lower_buttons_frame, width=20, textvariable=NEW_WORD)
    NEW_INPUT.grid(row=1, column=2, padx=20, pady=5)

    row = 2 #start at the beginning of the document

    left_button = ttk.Button(lower_buttons_frame, text="<<", state=DISABLED, command= lambda : backward_row(row))

    left_button.grid(row=1,column=0, padx=20)

    enter_button = ttk.Button(lower_buttons_frame, text="Enter", command= lambda: enter_word(NEW_WORD.get())) #insert cmd
    enter_button.grid(row=2, column=2, pady=10)

    right_button = ttk.Button(lower_buttons_frame, text=">>", command = lambda: forward_row(row+1))
    if(max_row == 2):
        right_button = ttk.Button(lower_buttons_frame, text=">>", state=DISABLED, command=lambda :forward_row(row+1))

    right_button.grid(row=1, column=4, padx=20)


    #Button to open excel

    open_excel_button = ttk.Button(lower_buttons_frame, text="Open Excel", command= lambda : open_excel())
    open_excel_button.grid(row=2, column=4, padx=20)

    #Button for the individual translation of each english word!

    individual_button = ttk.Button(lower_buttons_frame, text="Individual", command= lambda : translate_individual())
    individual_button.grid(row=2, column=0, padx=20)

    REVIEW_SCREEN.bind('<Return>', handler_enter_word)


#display_info
#Pre: The current window as a key
#Post: Depending on the key, will give different info regarding instructions for the current opened window
def display_info(box):
    if box == "main_info":
        messagebox.showinfo(message="Guess the word in English/Vietnamese based on an image and its word.", title="Random Words")
        messagebox.showinfo(message="Goes through all the words you have. You can add words here as your dictionary.", title="Review")
        messagebox.showinfo(message="Exit will prompt to choose whether to send your reviewed words to your email.", title="Exit")
    elif box == "random_info":
        messagebox.showinfo(message="Guess the word in English/Vietnamese based on an image and its word. The hint button will give you a definition as a clue.", title="Random Words")
    elif box == "review_info":
        messagebox.showinfo(message="Reviewing words takes from the local excel sheet created. Use the arrow buttons to traverse.", title="Review Words")
        messagebox.showinfo(message="Enter a word to create a new entry or revisit a word.", title="Review Words")
        messagebox.showinfo(message=r"The 'individual' button returns the individual translation of each word in the example and definition.", title="Review Words")
        messagebox.showinfo(message="Random button returns a random entry.", title="Review Words")
        messagebox.showinfo(message=r"'Excel' opens up the current excel word list.", title="Review Words")
#exit_out
#Pre: A ttk screen
#Post: Outputs a message box to ask the user to quit the window and close it out
def exit_out(screen):
    if messagebox.askquestion(message="Are you sure you want to quit?", title="Exit current tab") == "yes":
        screen.destroy()
        root.state("normal")

#main_screen
#Pre: None
#Post: Outputs the necessary widgets on the Main window (main menu)
def main_screen():
    global mainScreen
    global root
    root = Tk()

    global viet_label
    global define_label
    global example_label
    global define_label_v
    global example_label_v
    global eng_define_label
    global viet_define_label
    global eng_sent_label
    global viet_sent_label
    global word_label

    define_label = ttk.Label(root)
    example_label = ttk.Label(root)
    define_label_v = ttk.Label(root)
    example_label_v = ttk.Label(root)
    eng_define_label = ttk.Label(root)
    viet_define_label = ttk.Label(root)
    eng_sent_label = ttk.Label(root)
    viet_sent_label = ttk.Label(root)
    word_label = ttk.Label(root)
    viet_label = ttk.Label(root)

    initalize_list()

    root.title(" My Viet Dict ")
    root.geometry("250x350") #x&y
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    root.resizable(FALSE,FALSE)

    mainScreen = ttk.Frame(root, padding=(0,10,0,0)) #bot, top, right, left
    mainScreen.grid(row=0, column=0, sticky=(N, W, E, S))
    ttk.Label(mainScreen, padding=(5,5)).grid(row=0,column=0)

    learn = ttk.Label(mainScreen, relief="groove",text = "Learn a Random Word", font="Georgia 16", padding=(5,10))
    learn.grid(row=0, column=1, columnspan=1)
    rand = ttk.Button(mainScreen, text="Random Word", command = learn_random)
    rand.grid(row=1, column=1, columnspan=1)


    reviewL = ttk.Label(mainScreen, relief="groove", text = "Review Your Words", font="Georgia 16", padding=(5,10))
    reviewL.grid(row=2, column=1, columnspan=1)
    review =ttk.Button(mainScreen,text="Review", command = review_words)
    review.grid(row=3, column=1, columnspan=1 )

    exitL = ttk.Label(mainScreen, relief="groove", text = "Save and Exit", font="Georgia 16", padding=(5,10))
    exitL.grid(row=4, column=1, columnspan=1)
    exitB =ttk.Button(mainScreen,text="Exit", command= lambda: exit_out(root))
    exitB.grid(row=5, column=1, columnspan=1)

    infomationB = ttk.Button(mainScreen, text ="Info", command= lambda: display_info("main_info"))
    Label(mainScreen, text="").grid(row=6, column=1)
    infomationB.grid(row=7, column=1, columnspan=1)

    mainScreen.mainloop()

main_screen()

